import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime

class ExcelLogger:
    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Test Results"
        self.header_style = Font(bold=True)
        self._create_headers()

    def _create_headers(self):
        headers = ["Test Case", "Result", "Timestamp"]
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_style

    def log_result(self, test_case, result):
        row_num = self.ws.max_row + 1
        self.ws.cell(row=row_num, column=1, value=test_case)
        self.ws.cell(row=row_num, column=3, value=datetime.now())
        if result:
            self.ws.cell(row=row_num, column=2, value="Pass")
        else:
            self.ws.cell(row=row_num, column=2, value="Fail")
        self.wb.save(self.filename)

# Example usage:

